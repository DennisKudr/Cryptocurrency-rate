from PyQt5 import QtCore, QtGui, QtWidgets

import cryptocompare
import openpyxl
import os
from openpyxl import Workbook
import datetime
from datetime import datetime as dt
from pandas import DataFrame

from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
import sys
from design1 import Ui_MainWindow

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent = None, *args, **kwargs):
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.comboBox.activated.connect(self.getCurrency)
        self.pushButton.clicked.connect(self.getRate)
        self.listWidget_2.clicked.connect(self.Currency)
        self.pushButton_2.clicked.connect(self.saving)


    def getCurrency(self):
        global Currency
        Currencies = {'Etherium': 'ETH', 'Bitcoin': 'BTC', 'Litecoin': 'LTC'}
        Currency = Currencies[self.comboBox.currentText().strip()]

    def Currency(self):
        global curr
        curr = self.listWidget_2.currentItem().text()

    def plot(self):
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.plot(Date, Rate)
        return self.canvas.draw()

    def getDate(self):
        if self.checkBox.isChecked() is True:
            global Date
            Date = dt.now()
        if self.lineEdit_3.text():
            try:
                data = eval(self.lineEdit_3.text())
                Date = dt(data[0], data[1], data[2])
            except:
                pass
        if self.lineEdit.text() and self.lineEdit_2.text():
            try:
                start = eval(self.lineEdit.text())
                end = eval(self.lineEdit_2.text())
                start_date = dt(start[0], start[1], start[2])
                end_date = dt(end[0], end[1], end[2])
                delta = end_date - start_date
                Date = []
                if delta.days < 180:
                    for i in range(delta.days + 1):
                        day = start_date + datetime.timedelta(i)
                        Date.append(day)
                else:
                    self.textBrowser.setText('Choose less than 180 days period')
            except:
                pass
        return Date



    def getRate(self):
        global Date
        try:
            Date = self.getDate()
            global Rate
            if type(Date) == dt:
                try:
                    Rate = cryptocompare.get_historical_price(Currency, curr=curr, timestamp=Date)
                    self.figure.clear()
                    self.textBrowser.setText(
                        'The rate of {} is {} {} for {}'.format(self.comboBox.currentText().strip(), Rate[Currency][curr],
                                                                 curr, Date))
                except NameError:
                    try:
                        Rate = cryptocompare.get_historical_price('ETH', curr=curr, timestamp=Date)
                        self.figure.clear()
                        self.textBrowser.setText('The rate of Etherium is {} {} for {}'.format(Rate['ETH'][curr], curr, Date))
                    except NameError:
                        self.textBrowser.setText('Choose the currency!')
            else:
                try:
                    Rate = []
                    for i in Date:
                        rate = cryptocompare.get_historical_price(Currency, curr=curr, timestamp=i)
                        Rate.append(rate[Currency][curr])
                    self.textBrowser.setText('The values are: \n {}'.format(Rate))
                    self.plot()
                except NameError:
                    try:
                        for i in Date:
                            rate = cryptocompare.get_historical_price('ETH', curr=curr, timestamp=i)
                            Rate.append(rate['ETH'][curr])
                        self.textBrowser.setText('The values are: \n {}'.format(Rate))
                        self.plot()
                    except NameError:
                        self.textBrowser.setText('Choose the currency!')
        except NameError:
            self.textBrowser.setText('Choose the date!')


    def saving(self):
        if 'Rate' not in globals():
            self.textBrowser.setText('No result yet!')
        else:
            if type(Date) == dt:
                try:
                    result = {'Crypto_Currency': self.comboBox.currentText().strip(), 'Currency': curr,
                              'The rate': Rate[Currency][curr], 'Date': Date}
                except NameError:
                    result = {'Crypto_Currency': 'Etherium', 'Currency': curr,
                              'The rate': Rate['ETH'][curr], 'Date': Date}
                directory = QFileDialog.getExistingDirectory(self, 'Choose directory', os.getenv('HOME'))
                wb = Workbook()
                ws = wb.active
                for rec, i in zip(result, range(1,5,1)):
                   ws.cell(row=1, column=i).value = rec
                   ws.cell(row=2, column=i).value = result[rec]
                if directory:
                    wb.save(directory + '/The_Rate.xlsx')
            else:
                result = DataFrame({'Date': Date, 'Rate': Rate})
                directory = QFileDialog.getExistingDirectory(self, 'Choose directory', os.getenv('HOME'))
                if directory:
                    result.to_excel(directory + '/The_Rate.xlsx', index=False)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    main = MainWindow()
    main.show()
    sys.exit(app.exec_())

